import csv
import sys
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def get_csv_data(filepath):
    """Extract headers and all rows from CSV file."""
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            headers = [h.strip() for h in next(reader)]
            rows = list(reader)
            return headers, rows
    except FileNotFoundError:
        print(f"Error: File '{filepath}' not found.")
        sys.exit(1)
    except Exception as e:
        print(f"Error reading '{filepath}': {e}")
        sys.exit(1)

def auto_detect_key_column(headers):
    """Try to auto-detect common ID column names."""
    common_ids = ['sys_id', 'id', 'ID', 'number', 'Number', 'record_id', 
                  'recordid', 'ticket_number', 'user_id', 'userid']
    
    for id_col in common_ids:
        if id_col in headers:
            return id_col
    return None

def prompt_for_key_column(api_headers, frontend_headers):
    """Ask user to select key column if auto-detect fails."""
    common_cols = set(api_headers) & set(frontend_headers)
    
    print("\n" + "="*70)
    print("KEY COLUMN SELECTION")
    print("="*70)
    print("Could not auto-detect a key column. Please select one from common columns:")
    print()
    
    common_list = sorted(common_cols)
    for i, col in enumerate(common_list, 1):
        print(f"{i}. {col}")
    
    while True:
        try:
            choice = input("\nEnter column number (or column name): ").strip()
            if choice.isdigit():
                idx = int(choice) - 1
                if 0 <= idx < len(common_list):
                    return common_list[idx]
            elif choice in common_cols:
                return choice
            print("Invalid choice. Please try again.")
        except KeyboardInterrupt:
            print("\nOperation cancelled.")
            sys.exit(0)

def create_row_dict(headers, rows, key_col):
    """Create dictionary with key column as key."""
    key_idx = headers.index(key_col)
    row_dict = {}
    for row in rows:
        if len(row) > key_idx:
            key_val = row[key_idx].strip()
            if key_val:  # Skip empty keys
                row_dict[key_val] = row
    return row_dict

def compare_csv_files(api_file, frontend_file, output_excel):
    """Main comparison function."""
    print("\n" + "="*70)
    print("CSV COMPARISON: API vs ServiceNow Frontend")
    print("="*70)
    
    # Load data
    api_headers, api_rows = get_csv_data(api_file)
    frontend_headers, frontend_rows = get_csv_data(frontend_file)
    
    print(f"\nAPI CSV: {len(api_rows)} rows, {len(api_headers)} columns")
    print(f"Frontend CSV: {len(frontend_rows)} rows, {len(frontend_headers)} columns")
    
    # Detect or prompt for key column
    key_col = auto_detect_key_column(api_headers)
    
    if key_col and key_col in frontend_headers:
        print(f"\n✓ Auto-detected key column: '{key_col}'")
    else:
        key_col = prompt_for_key_column(api_headers, frontend_headers)
        print(f"\n✓ Using key column: '{key_col}'")
    
    # Compare columns
    api_col_set = set(api_headers)
    frontend_col_set = set(frontend_headers)
    common_columns = api_col_set & frontend_col_set
    missing_in_api = frontend_col_set - api_col_set
    extra_in_api = api_col_set - frontend_col_set
    
    # Create row dictionaries
    api_dict = create_row_dict(api_headers, api_rows, key_col)
    frontend_dict = create_row_dict(frontend_headers, frontend_rows, key_col)
    
    # Find missing rows
    api_keys = set(api_dict.keys())
    frontend_keys = set(frontend_dict.keys())
    only_in_api = api_keys - frontend_keys
    only_in_frontend = frontend_keys - api_keys
    common_keys = api_keys & frontend_keys
    
    # Compare values for common rows
    value_differences = []
    for key in common_keys:
        api_row = api_dict[key]
        frontend_row = frontend_dict[key]
        
        for col in common_columns:
            if col == key_col:
                continue
            
            api_idx = api_headers.index(col)
            frontend_idx = frontend_headers.index(col)
            
            api_val = api_row[api_idx].strip() if api_idx < len(api_row) else ""
            frontend_val = frontend_row[frontend_idx].strip() if frontend_idx < len(frontend_row) else ""
            
            if api_val != frontend_val:
                value_differences.append({
                    'key': key,
                    'column': col,
                    'api_value': api_val,
                    'frontend_value': frontend_val
                })
    
    # Print summary
    print("\n" + "="*70)
    print("COMPARISON SUMMARY")
    print("="*70)
    print(f"✓ Common columns: {len(common_columns)}")
    print(f"⚠ Missing in API: {len(missing_in_api)}")
    print(f"⚡ Extra in API: {len(extra_in_api)}")
    print(f"\n✓ Rows in both files: {len(common_keys)}")
    print(f"⚠ Rows only in API: {len(only_in_api)}")
    print(f"⚠ Rows only in Frontend: {len(only_in_frontend)}")
    print(f"\n⚠ Value differences found: {len(value_differences)}")
    
    # Create Excel report
    create_excel_report(
        output_excel, key_col, api_headers, frontend_headers,
        common_columns, missing_in_api, extra_in_api,
        only_in_api, only_in_frontend, api_dict, frontend_dict,
        value_differences
    )
    
    print(f"\n✓ Excel report created: {output_excel}")
    print("="*70)

def create_excel_report(filename, key_col, api_headers, frontend_headers,
                       common_columns, missing_in_api, extra_in_api,
                       only_in_api, only_in_frontend, api_dict, frontend_dict,
                       value_differences):
    """Create Excel workbook with comparison results."""
    wb = openpyxl.Workbook()
    
    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    warning_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    summary_data = [
        ["CSV Comparison Report", ""],
        ["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Key Column Used", key_col],
        ["", ""],
        ["COLUMN COMPARISON", ""],
        ["Total API Columns", len(api_headers)],
        ["Total Frontend Columns", len(frontend_headers)],
        ["Common Columns", len(common_columns)],
        ["Missing in API", len(missing_in_api)],
        ["Extra in API", len(extra_in_api)],
        ["", ""],
        ["ROW COMPARISON", ""],
        ["Total API Rows", len(api_dict)],
        ["Total Frontend Rows", len(frontend_dict)],
        ["Rows in Both", len(set(api_dict.keys()) & set(frontend_dict.keys()))],
        ["Rows Only in API", len(only_in_api)],
        ["Rows Only in Frontend", len(only_in_frontend)],
        ["", ""],
        ["VALUE COMPARISON", ""],
        ["Value Differences Found", len(value_differences)],
    ]
    
    for row_data in summary_data:
        ws_summary.append(row_data)
    
    # Format summary
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 20
    for row in [1, 5, 12, 19]:
        ws_summary[f'A{row}'].font = Font(bold=True, size=12)
    
    # Sheet 2: Column Comparison
    ws_columns = wb.create_sheet("Column Comparison")
    ws_columns.append(["Status", "Column Name"])
    
    for col in ws_columns[1]:
        col.fill = header_fill
        col.font = header_font
    
    for col in sorted(common_columns):
        ws_columns.append(["✓ Common", col])
    for col in sorted(missing_in_api):
        row = ws_columns.max_row + 1
        ws_columns.append(["⚠ Missing in API", col])
        ws_columns[f'A{row}'].fill = warning_fill
    for col in sorted(extra_in_api):
        ws_columns.append(["⚡ Extra in API", col])
    
    ws_columns.column_dimensions['A'].width = 20
    ws_columns.column_dimensions['B'].width = 40
    
    # Sheet 3: Missing Rows
    ws_missing = wb.create_sheet("Missing Rows")
    ws_missing.append(["Status", key_col, "Additional Info"])
    
    for col in ws_missing[1]:
        col.fill = header_fill
        col.font = header_font
    
    for key in sorted(only_in_api):
        row_data = api_dict[key]
        info = " | ".join([f"{api_headers[i]}: {row_data[i]}" for i in range(min(3, len(row_data))) if i < len(api_headers)])
        ws_missing.append(["Only in API", key, info])
    
    for key in sorted(only_in_frontend):
        row_data = frontend_dict[key]
        info = " | ".join([f"{frontend_headers[i]}: {row_data[i]}" for i in range(min(3, len(row_data))) if i < len(frontend_headers)])
        row_num = ws_missing.max_row + 1
        ws_missing.append(["Only in Frontend", key, info])
        ws_missing[f'A{row_num}'].fill = warning_fill
    
    ws_missing.column_dimensions['A'].width = 20
    ws_missing.column_dimensions['B'].width = 30
    ws_missing.column_dimensions['C'].width = 60
    
    # Sheet 4: Value Differences
    ws_diff = wb.create_sheet("Value Differences")
    ws_diff.append([key_col, "Column", "API Value", "Frontend Value"])
    
    for col in ws_diff[1]:
        col.fill = header_fill
        col.font = header_font
    
    for diff in value_differences:
        ws_diff.append([diff['key'], diff['column'], diff['api_value'], diff['frontend_value']])
    
    ws_diff.column_dimensions['A'].width = 25
    ws_diff.column_dimensions['B'].width = 25
    ws_diff.column_dimensions['C'].width = 30
    ws_diff.column_dimensions['D'].width = 30
    
    wb.save(filename)

if __name__ == "__main__":
    # Update these paths
    API_CSV = "api_data.csv"
    FRONTEND_CSV = "frontend_columns.csv"
    OUTPUT_EXCEL = f"comparison_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    print("\nStarting CSV comparison...")
    print(f"API CSV: {API_CSV}")
    print(f"Frontend CSV: {FRONTEND_CSV}")
    print(f"Output: {OUTPUT_EXCEL}")
    
    compare_csv_files(API_CSV, FRONTEND_CSV, OUTPUT_EXCEL)
