import csv
import sys
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def get_csv_data(filepath):
    """Extract headers and all rows from CSV file."""
    encodings = ['utf-8', 'latin-1', 'iso-8859-1', 'cp1252', 'utf-16']
    
    for encoding in encodings:
        try:
            with open(filepath, 'r', encoding=encoding) as f:
                reader = csv.reader(f)
                headers = [h.strip() for h in next(reader)]
                rows = list(reader)
                print(f"Successfully read '{filepath}' with {encoding} encoding")
                return headers, rows
        except UnicodeDecodeError:
            continue
        except FileNotFoundError:
            print(f"Error: File '{filepath}' not found.")
            sys.exit(1)
        except Exception as e:
            if encoding == encodings[-1]:  # Last encoding attempt
                print(f"Error reading '{filepath}': {e}")
                sys.exit(1)
    
    print(f"Error: Could not read '{filepath}' with any supported encoding.")
    sys.exit(1)

def auto_detect_key_column(headers):
    """Try to auto-detect common ID column names."""
    common_ids = ['sys_id', 'id', 'ID', 'number', 'Number', 'record_id', 
                  'recordid', 'ticket_number', 'user_id', 'userid', 'email', 
                  'employee_id', 'employeeid']
    
    for id_col in common_ids:
        if id_col in headers:
            return id_col
    return None

def prompt_for_key_column(prod_headers, api_headers):
    """Ask user to select key column if auto-detect fails."""
    common_cols = set(prod_headers) & set(api_headers)
    
    print("\n" + "="*70)
    print("KEY COLUMN SELECTION")
    print("="*70)
    print("Could not auto-detect a key column. Please select one from common columns:")
    print("(This should be a column that uniquely identifies each row in PROD)")
    print()
    
    common_list = sorted(common_cols)
    for i, col in enumerate(common_list, 1):
        print(f"{i}. {col}")
    
    while True:
        try:
            choice = input("\nEnter column number (or column name): ").strip()
            if choice.isdigit():
                idx = int(choice) - 1
                if 0 <= idx < len(common_list):
                    return common_list[idx]
            elif choice in common_cols:
                return choice
            print("Invalid choice. Please try again.")
        except KeyboardInterrupt:
            print("\nOperation cancelled.")
            sys.exit(0)

def create_row_dict(headers, rows, key_col):
    """Create dictionary with key column as key."""
    key_idx = headers.index(key_col)
    row_dict = {}
    rows_with_empty_keys = []
    
    for row_num, row in enumerate(rows, start=2):  # start=2 because row 1 is headers
        if len(row) > key_idx:
            key_val = row[key_idx].strip()
            if key_val:  # Only add rows with non-empty keys
                row_dict[key_val] = {'data': row, 'row_num': row_num}
            else:
                rows_with_empty_keys.append({'row_num': row_num, 'data': row})
        else:
            rows_with_empty_keys.append({'row_num': row_num, 'data': row})
    
    return row_dict, rows_with_empty_keys

def compare_csv_files(prod_file, api_file, output_excel):
    """Main comparison function - PROD is master source."""
    print("\n" + "="*70)
    print("CSV COMPARISON: API Validation Against PROD (Master Source)")
    print("="*70)
    
    # Load data - PROD first as it's the master
    prod_headers, prod_rows = get_csv_data(prod_file)
    api_headers, api_rows = get_csv_data(api_file)
    
    print(f"\nPROD (Master): {len(prod_rows)} rows, {len(prod_headers)} columns")
    print(f"API (Validate): {len(api_rows)} rows, {len(api_headers)} columns")
    
    # Detect or prompt for key column (prioritize PROD headers)
    key_col = auto_detect_key_column(prod_headers)
    
    if key_col and key_col in api_headers:
        print(f"\n✓ Auto-detected key column: '{key_col}'")
    else:
        key_col = prompt_for_key_column(prod_headers, api_headers)
        print(f"\n✓ Using key column: '{key_col}'")
    
    # Compare columns
    prod_col_set = set(prod_headers)
    api_col_set = set(api_headers)
    common_columns = prod_col_set & api_col_set
    missing_in_api = prod_col_set - api_col_set
    extra_in_api = api_col_set - prod_col_set
    
    # Create row dictionaries
    prod_dict, prod_empty_keys = create_row_dict(prod_headers, prod_rows, key_col)
    api_dict, api_empty_keys = create_row_dict(api_headers, api_rows, key_col)
    
    # Find missing rows (using PROD as baseline)
    prod_keys = set(prod_dict.keys())
    api_keys = set(api_dict.keys())
    missing_from_api = prod_keys - api_keys  # In PROD but not in API
    extra_in_api_keys = api_keys - prod_keys  # In API but not in PROD
    common_keys = prod_keys & api_keys
    
    # Compare values for common rows (using PROD as source of truth)
    value_differences = []
    matching_rows = []
    
    for key in common_keys:
        prod_row = prod_dict[key]['data']
        api_row = api_dict[key]['data']
        has_difference = False
        row_differences = []
        
        for col in common_columns:
            if col == key_col:
                continue
            
            prod_idx = prod_headers.index(col)
            api_idx = api_headers.index(col)
            
            prod_val = prod_row[prod_idx].strip() if prod_idx < len(prod_row) else ""
            api_val = api_row[api_idx].strip() if api_idx < len(api_row) else ""
            
            if api_val != prod_val:
                has_difference = True
                row_differences.append({
                    'key': key,
                    'column': col,
                    'prod_value': prod_val,
                    'api_value': api_val,
                    'issue': 'Missing' if not api_val and prod_val else 'Different'
                })
        
        if has_difference:
            value_differences.extend(row_differences)
        else:
            matching_rows.append(key)
    
    # Calculate API data quality metrics
    total_prod_rows = len(prod_dict)
    api_completeness = (len(common_keys) / total_prod_rows * 100) if total_prod_rows > 0 else 0
    api_accuracy = (len(matching_rows) / len(common_keys) * 100) if common_keys else 0
    
    # Print summary
    print("\n" + "="*70)
    print("VALIDATION SUMMARY")
    print("="*70)
    print(f"\n📊 COLUMN COMPARISON:")
    print(f"   ✓ Common columns: {len(common_columns)}")
    print(f"   ⚠ Missing in API (present in PROD): {len(missing_in_api)}")
    print(f"   ⚡ Extra in API (not in PROD): {len(extra_in_api)}")
    
    print(f"\n📊 ROW COMPARISON (PROD as baseline):")
    print(f"   Total PROD rows: {total_prod_rows}")
    print(f"   ✓ Found in API: {len(common_keys)}")
    print(f"   ⚠ Missing from API: {len(missing_from_api)}")
    print(f"   ⚡ Extra in API: {len(extra_in_api_keys)}")
    
    print(f"\n📊 DATA QUALITY:")
    print(f"   ✓ Rows matching PROD exactly: {len(matching_rows)}")
    print(f"   ⚠ Rows with differences: {len(common_keys) - len(matching_rows)}")
    print(f"   ⚠ Total value mismatches: {len(value_differences)}")
    print(f"   ⚠ API rows with empty keys: {len(api_empty_keys)}")
    if prod_empty_keys:
        print(f"   ⚠ PROD rows with empty keys: {len(prod_empty_keys)}")
    
    print(f"\n📊 API COMPLETENESS: {api_completeness:.1f}% of PROD data")
    print(f"📊 API ACCURACY: {api_accuracy:.1f}% of matched rows are correct")
    
    # Create Excel report
    create_excel_report(
        output_excel, key_col, prod_headers, api_headers,
        common_columns, missing_in_api, extra_in_api,
        missing_from_api, extra_in_api_keys, prod_dict, api_dict,
        value_differences, matching_rows, api_empty_keys, prod_empty_keys,
        api_completeness, api_accuracy
    )
    
    print(f"\n✓ Excel validation report created: {output_excel}")
    print("="*70)

def create_excel_report(filename, key_col, prod_headers, api_headers,
                       common_columns, missing_in_api, extra_in_api,
                       missing_from_api, extra_in_api_keys, prod_dict, api_dict,
                       value_differences, matching_rows, api_empty_keys, prod_empty_keys,
                       api_completeness, api_accuracy):
    """Create Excel workbook with validation results."""
    wb = openpyxl.Workbook()
    
    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    summary_data = [
        ["API VALIDATION REPORT - PROD as Master Source", ""],
        ["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Key Column Used", key_col],
        ["", ""],
        ["COLUMN COMPARISON", ""],
        ["Total PROD Columns", len(prod_headers)],
        ["Total API Columns", len(api_headers)],
        ["Common Columns", len(common_columns)],
        ["Missing in API", len(missing_in_api)],
        ["Extra in API", len(extra_in_api)],
        ["", ""],
        ["ROW COMPARISON (PROD Baseline)", ""],
        ["Total PROD Rows", len(prod_dict)],
        ["Total API Rows", len(api_dict)],
        ["Rows Found in API", len(set(prod_dict.keys()) & set(api_dict.keys()))],
        ["Rows Missing from API", len(missing_from_api)],
        ["Extra Rows in API", len(extra_in_api_keys)],
        ["", ""],
        ["DATA QUALITY METRICS", ""],
        ["Rows Matching PROD Exactly", len(matching_rows)],
        ["Rows with Value Differences", len(set(prod_dict.keys()) & set(api_dict.keys())) - len(matching_rows)],
        ["Total Value Mismatches", len(value_differences)],
        ["API Rows with Empty Keys", len(api_empty_keys)],
        ["PROD Rows with Empty Keys", len(prod_empty_keys)],
        ["", ""],
        ["API COMPLETENESS SCORE", f"{api_completeness:.1f}%"],
        ["API ACCURACY SCORE", f"{api_accuracy:.1f}%"],
    ]
    
    for row_data in summary_data:
        ws_summary.append(row_data)
    
    # Format summary
    ws_summary.column_dimensions['A'].width = 35
    ws_summary.column_dimensions['B'].width = 25
    for row in [1, 5, 12, 19, 26]:
        ws_summary[f'A{row}'].font = Font(bold=True, size=12)
    
    # Color code quality scores
    completeness_row = 26
    accuracy_row = 27
    if api_completeness >= 95:
        ws_summary[f'B{completeness_row}'].fill = success_fill
    elif api_completeness >= 80:
        ws_summary[f'B{completeness_row}'].fill = warning_fill
    else:
        ws_summary[f'B{completeness_row}'].fill = error_fill
    
    if api_accuracy >= 95:
        ws_summary[f'B{accuracy_row}'].fill = success_fill
    elif api_accuracy >= 80:
        ws_summary[f'B{accuracy_row}'].fill = warning_fill
    else:
        ws_summary[f'B{accuracy_row}'].fill = error_fill
    
    # Sheet 2: Column Comparison
    ws_columns = wb.create_sheet("Column Comparison")
    ws_columns.append(["Status", "Column Name", "Notes"])
    
    for col in ws_columns[1]:
        col.fill = header_fill
        col.font = header_font
    
    for col in sorted(common_columns):
        ws_columns.append(["✓ Common", col, "Present in both"])
    for col in sorted(missing_in_api):
        row = ws_columns.max_row + 1
        ws_columns.append(["⚠ MISSING IN API", col, "Required from PROD"])
        ws_columns[f'A{row}'].fill = error_fill
        ws_columns[f'B{row}'].fill = error_fill
    for col in sorted(extra_in_api):
        ws_columns.append(["⚡ Extra in API", col, "Not in PROD"])
    
    ws_columns.column_dimensions['A'].width = 20
    ws_columns.column_dimensions['B'].width = 40
    ws_columns.column_dimensions['C'].width = 30
    
    # Sheet 3: Missing Rows from API
    ws_missing = wb.create_sheet("Missing from API")
    ws_missing.append([key_col, "Status", "PROD Data Preview"])
    
    for col in ws_missing[1]:
        col.fill = header_fill
        col.font = header_font
    
    for key in sorted(missing_from_api):
        row_data = prod_dict[key]['data']
        info = " | ".join([f"{prod_headers[i]}: {row_data[i]}" for i in range(min(4, len(row_data))) if i < len(prod_headers)])
        row_num = ws_missing.max_row + 1
        ws_missing.append([key, "⚠ MISSING", info])
        ws_missing[f'A{row_num}'].fill = error_fill
        ws_missing[f'B{row_num}'].fill = error_fill
    
    ws_missing.column_dimensions['A'].width = 30
    ws_missing.column_dimensions['B'].width = 15
    ws_missing.column_dimensions['C'].width = 80
    
    # Sheet 4: Value Differences
    ws_diff = wb.create_sheet("Value Differences")
    ws_diff.append([key_col, "Column", "Issue Type", "PROD Value (Correct)", "API Value (Incorrect)"])
    
    for col in ws_diff[1]:
        col.fill = header_fill
        col.font = header_font
    
    for diff in value_differences:
        row_num = ws_diff.max_row + 1
        ws_diff.append([
            diff['key'], 
            diff['column'], 
            diff['issue'],
            diff['prod_value'], 
            diff['api_value']
        ])
        if diff['issue'] == 'Missing':
            ws_diff[f'C{row_num}'].fill = error_fill
            ws_diff[f'E{row_num}'].fill = error_fill
        else:
            ws_diff[f'C{row_num}'].fill = warning_fill
    
    ws_diff.column_dimensions['A'].width = 25
    ws_diff.column_dimensions['B'].width = 25
    ws_diff.column_dimensions['C'].width = 15
    ws_diff.column_dimensions['D'].width = 35
    ws_diff.column_dimensions['E'].width = 35
    
    # Sheet 5: Extra in API
    ws_extra = wb.create_sheet("Extra in API")
    ws_extra.append([key_col, "Status", "API Data Preview", "Notes"])
    
    for col in ws_extra[1]:
        col.fill = header_fill
        col.font = header_font
    
    for key in sorted(extra_in_api_keys):
        row_data = api_dict[key]['data']
        info = " | ".join([f"{api_headers[i]}: {row_data[i]}" for i in range(min(4, len(row_data))) if i < len(api_headers)])
        ws_extra.append([key, "Extra", info, "Not in PROD - may be new or error"])
    
    ws_extra.column_dimensions['A'].width = 30
    ws_extra.column_dimensions['B'].width = 15
    ws_extra.column_dimensions['C'].width = 80
    ws_extra.column_dimensions['D'].width = 30
    
    # Sheet 6: API Data Issues
    if api_empty_keys or prod_empty_keys:
        ws_issues = wb.create_sheet("Data Issues")
        ws_issues.append(["Issue Type", "Source", "Row Number", "Data Preview"])
        
        for col in ws_issues[1]:
            col.fill = header_fill
            col.font = header_font
        
        for item in api_empty_keys:
            info = " | ".join([f"{api_headers[i]}: {item['data'][i]}" for i in range(min(4, len(item['data']))) if i < len(api_headers)])
            row_num = ws_issues.max_row + 1
            ws_issues.append([f"Empty {key_col}", "API", item['row_num'], info])
            ws_issues[f'A{row_num}'].fill = error_fill
        
        for item in prod_empty_keys:
            info = " | ".join([f"{prod_headers[i]}: {item['data'][i]}" for i in range(min(4, len(item['data']))) if i < len(prod_headers)])
            row_num = ws_issues.max_row + 1
            ws_issues.append([f"Empty {key_col}", "PROD", item['row_num'], info])
            ws_issues[f'A{row_num}'].fill = warning_fill
        
        ws_issues.column_dimensions['A'].width = 20
        ws_issues.column_dimensions['B'].width = 15
        ws_issues.column_dimensions['C'].width = 15
        ws_issues.column_dimensions['D'].width = 80
    
    wb.save(filename)

if __name__ == "__main__":
    # Update these paths
    PROD_CSV = "prod_data.csv"  # Master source of truth
    API_CSV = "api_data.csv"    # Data to validate
    OUTPUT_EXCEL = f"api_validation_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    print("\nStarting API validation against PROD (Master Source)...")
    print(f"PROD CSV (Master): {PROD_CSV}")
    print(f"API CSV (Validate): {API_CSV}")
    print(f"Output: {OUTPUT_EXCEL}")
    
    compare_csv_files(PROD_CSV, API_CSV, OUTPUT_EXCEL)
