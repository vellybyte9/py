"""
CSV COMPARISON SCRIPT - DETAILED EXPLANATION
============================================
Purpose: Compare API data against PROD (master source) to validate data quality
Author: Created for data validation workflow
"""

# ============================================================================
# SECTION 1: IMPORT LIBRARIES
# ============================================================================
# These are pre-built Python tools we need for our script

import csv
# csv = helps us read and write CSV (comma-separated values) files
# Example: turns "John,Doe,30" into ["John", "Doe", "30"]

import sys
# sys = system functions, helps us exit the program when there's an error
# Example: sys.exit(1) stops the program

from datetime import datetime
# datetime = works with dates and times
# Example: datetime.now() gives us current date/time like "2025-01-02 14:30:22"

import openpyxl
# openpyxl = library to create and edit Excel files (.xlsx)
# This is what lets us make the fancy Excel report

from openpyxl.styles import Font, PatternFill, Alignment
# These are styling tools from openpyxl
# Font = change text style (bold, color, size)
# PatternFill = add background colors to cells (red, yellow, green)
# Alignment = align text (left, center, right)

from openpyxl.utils import get_column_letter
# get_column_letter = converts numbers to Excel column letters
# Example: 1 becomes "A", 2 becomes "B", 27 becomes "AA"


# ============================================================================
# SECTION 2: FUNCTION TO READ CSV FILES
# ============================================================================

def get_csv_data(filepath):
    """
    Extract headers and all rows from CSV file.
    
    What this does:
    - Opens a CSV file
    - Reads the first row (headers/column names)
    - Reads all remaining rows (data)
    - Returns both headers and data
    
    Parameters:
    - filepath: string, the path to the CSV file (like "C:/data/file.csv")
    
    Returns:
    - headers: list of column names ["name", "email", "age"]
    - rows: list of lists, each inner list is one row of data
    """
    
    # List of different text encodings to try
    # Encoding = how text is stored in the file (like different languages)
    encodings = ['utf-8', 'latin-1', 'iso-8859-1', 'cp1252', 'utf-16']
    # utf-8 = most common, works for English and international characters
    # latin-1, iso-8859-1, cp1252 = Windows/Western European encodings
    # utf-16 = another Unicode format
    
    # Try each encoding until one works
    for encoding in encodings:
        try:
            # Try to open the file with current encoding
            with open(filepath, 'r', encoding=encoding) as f:
                # 'r' = read mode (don't change the file)
                # 'with' = automatically closes file when done
                # 'as f' = call the opened file 'f'
                
                reader = csv.reader(f)
                # csv.reader = tool that reads CSV files line by line
                
                headers = [h.strip() for h in next(reader)]
                # next(reader) = get the first line (headers)
                # h.strip() = remove extra spaces from each header
                # [h.strip() for h in ...] = list comprehension, does strip() on each item
                # Result: ["name", "email", "age"] with no extra spaces
                
                rows = list(reader)
                # list(reader) = read all remaining lines into a list
                # Result: [["John", "john@email.com", "30"], ["Jane", "jane@email.com", "25"], ...]
                
                print(f"Successfully read '{filepath}' with {encoding} encoding")
                # f"..." = f-string, lets us insert variables into text
                # Prints: "Successfully read 'data.csv' with utf-8 encoding"
                
                return headers, rows
                # Return both headers and rows to whoever called this function
                
        except UnicodeDecodeError:
            # This error happens when encoding is wrong (can't read the text)
            continue
            # continue = skip to next encoding in the loop
            
        except FileNotFoundError:
            # This error happens when the file doesn't exist
            print(f"Error: File '{filepath}' not found.")
            sys.exit(1)
            # sys.exit(1) = stop the program with error code 1
            
        except Exception as e:
            # This catches any other unexpected errors
            # 'as e' stores the error message in variable 'e'
            if encoding == encodings[-1]:  # Last encoding attempt
                # encodings[-1] = last item in encodings list
                # Only print error if we tried all encodings
                print(f"Error reading '{filepath}': {e}")
                sys.exit(1)
    
    # If we get here, none of the encodings worked
    print(f"Error: Could not read '{filepath}' with any supported encoding.")
    sys.exit(1)


# ============================================================================
# SECTION 3: AUTO-DETECT KEY COLUMN
# ============================================================================

def auto_detect_key_column(headers):
    """
    Try to auto-detect common ID column names.
    
    What this does:
    - Looks through column headers for common ID field names
    - Returns the first match it finds
    - Returns None if no common ID is found
    
    Parameters:
    - headers: list of column names from the CSV
    
    Returns:
    - String with column name if found (like "sys_id")
    - None if no common ID column found
    """
    
    # List of common ID column names we should look for
    common_ids = ['sys_id', 'id', 'ID', 'number', 'Number', 'record_id', 
                  'recordid', 'ticket_number', 'user_id', 'userid', 'email', 
                  'employee_id', 'employeeid']
    # sys_id = ServiceNow's standard ID field
    # id/ID = generic ID fields
    # number = ticket numbers like INC0012345
    # email = can be unique identifier for users
    
    # Loop through each potential ID name
    for id_col in common_ids:
        if id_col in headers:
            # Check if this ID name exists in the CSV headers
            return id_col
            # Return it immediately if found (stops the loop)
    
    # If we get here, we didn't find any common ID
    return None
    # None = special Python value meaning "nothing found"


# ============================================================================
# SECTION 4: PROMPT USER TO SELECT KEY COLUMN
# ============================================================================

def prompt_for_key_column(prod_headers, api_headers):
    """
    Ask user to select key column if auto-detect fails.
    
    What this does:
    - Shows user all columns that exist in BOTH files
    - Lets user pick which one to use as the key
    - Validates the user's choice
    
    Parameters:
    - prod_headers: list of PROD column names
    - api_headers: list of API column names
    
    Returns:
    - String with the chosen column name
    """
    
    common_cols = set(prod_headers) & set(api_headers)
    # set() = convert list to set (removes duplicates, allows set operations)
    # & = intersection operator, finds items in BOTH sets
    # Result: only columns that exist in both PROD and API
    # Example: {"name", "email", "sys_id"}
    
    # Print header for this section
    print("\n" + "="*70)
    # "\n" = new line (blank line before the header)
    # "="*70 = repeat "=" 70 times to make a line
    print("KEY COLUMN SELECTION")
    print("="*70)
    print("Could not auto-detect a key column. Please select one from common columns:")
    print("(This should be a column that uniquely identifies each row in PROD)")
    print()  # blank line
    
    common_list = sorted(common_cols)
    # sorted() = put the columns in alphabetical order
    # We convert set to list because sets have no order
    
    # Show numbered list of columns
    for i, col in enumerate(common_list, 1):
        # enumerate(list, 1) = loop with counter starting at 1
        # i = the number (1, 2, 3...)
        # col = the column name
        print(f"{i}. {col}")
        # Prints: "1. email"
        #         "2. name"
        #         "3. sys_id"
    
    # Keep asking until we get valid input
    while True:
        # while True = loop forever until we use 'return' or 'break'
        
        try:
            # try = attempt this code, catch errors if they happen
            
            choice = input("\nEnter column number (or column name): ").strip()
            # input() = pause and wait for user to type something
            # .strip() = remove extra spaces from what they typed
            
            if choice.isdigit():
                # .isdigit() = check if the input is a number
                # True if they typed "3", False if they typed "email"
                
                idx = int(choice) - 1
                # int() = convert text to number
                # - 1 because lists start at 0 but we showed numbers starting at 1
                # If they typed "3", idx becomes 2
                
                if 0 <= idx < len(common_list):
                    # Check if the number is valid (not too low or too high)
                    # 0 <= idx means: idx is 0 or higher
                    # idx < len(common_list) means: idx is less than list length
                    return common_list[idx]
                    # Return the column at that position
                    # common_list[2] might be "sys_id"
                    
            elif choice in common_cols:
                # Check if they typed an exact column name
                return choice
                # Return what they typed
                
            # If we get here, their choice was invalid
            print("Invalid choice. Please try again.")
            # Loop continues, asks them again
            
        except KeyboardInterrupt:
            # KeyboardInterrupt = user pressed Ctrl+C to cancel
            print("\nOperation cancelled.")
            sys.exit(0)
            # sys.exit(0) = exit with success code (user chose to quit)


# ============================================================================
# SECTION 5: CREATE ROW DICTIONARY
# ============================================================================

def create_row_dict(headers, rows, key_col):
    """
    Create dictionary with key column as key.
    
    What this does:
    - Organizes rows by their unique ID for easy lookup
    - Tracks rows that have empty/missing ID values
    
    Parameters:
    - headers: list of column names
    - rows: list of data rows
    - key_col: which column to use as the key (like "sys_id")
    
    Returns:
    - row_dict: dictionary where key=ID, value=row data
    - rows_with_empty_keys: list of rows that have no ID
    
    Example:
    Input rows: [["123", "John"], ["456", "Jane"], ["", "Bob"]]
    Output row_dict: {"123": {"data": ["123", "John"], "row_num": 2},
                      "456": {"data": ["456", "Jane"], "row_num": 3}}
    Output empty: [{"row_num": 4, "data": ["", "Bob"]}]
    """
    
    key_idx = headers.index(key_col)
    # .index() = find position of key_col in headers list
    # If key_col is "sys_id" and headers is ["name", "sys_id", "email"]
    # Then key_idx = 1 (second position, counting from 0)
    
    row_dict = {}
    # {} = empty dictionary
    # Dictionary = stores key-value pairs like a phonebook
    # Example: {"John": "555-1234", "Jane": "555-5678"}
    
    rows_with_empty_keys = []
    # [] = empty list to store problem rows
    
    # Loop through each row with a counter
    for row_num, row in enumerate(rows, start=2):
        # enumerate(rows, start=2) = loop with counter starting at 2
        # start=2 because row 1 is headers in the CSV file
        # row_num = the CSV row number (2, 3, 4...)
        # row = the actual data ["123", "John", "john@email.com"]
        
        if len(row) > key_idx:
            # len(row) = how many columns in this row
            # Check if this row has enough columns to contain the key
            # Prevents error if a row is missing columns
            
            key_val = row[key_idx].strip()
            # row[key_idx] = get the value from the key column
            # If key_idx is 1 and row is ["John", "123", "john@email.com"]
            # Then key_val = "123"
            # .strip() = remove spaces
            
            if key_val:  # Only add rows with non-empty keys
                # if key_val = check if it's not empty
                # Empty string "" is False in Python
                # Any text like "123" is True
                
                row_dict[key_val] = {'data': row, 'row_num': row_num}
                # Add to dictionary
                # row_dict["123"] = {'data': ["John", "123", "john@..."], 'row_num': 2}
                # Now we can look up row by ID: row_dict["123"]
                
            else:
                # Key value is empty, this is a problem row
                rows_with_empty_keys.append({'row_num': row_num, 'data': row})
                # .append() = add to end of list
                # Store which row had the problem and what data it had
                
        else:
            # Row doesn't have enough columns, also a problem
            rows_with_empty_keys.append({'row_num': row_num, 'data': row})
    
    return row_dict, rows_with_empty_keys
    # Return both the organized dictionary and problem rows


# ============================================================================
# SECTION 6: MAIN COMPARISON FUNCTION
# ============================================================================

def compare_csv_files(prod_file, api_file, output_excel):
    """
    Main comparison function - PROD is master source.
    
    What this does:
    - Reads both CSV files
    - Finds the key column to match rows
    - Compares columns and values
    - Creates Excel report with findings
    
    This is the "brain" that coordinates everything.
    
    Parameters:
    - prod_file: path to PROD CSV (master data)
    - api_file: path to API CSV (data to validate)
    - output_excel: where to save the Excel report
    """
    
    # Print header
    print("\n" + "="*70)
    print("CSV COMPARISON: API Validation Against PROD (Master Source)")
    print("="*70)
    
    # Load data - PROD first as it's the master
    prod_headers, prod_rows = get_csv_data(prod_file)
    # Call our function from Section 2
    # Returns two things: headers and rows
    # prod_headers = ["sys_id", "name", "email"]
    # prod_rows = [["123", "John", "john@..."], ["456", "Jane", "jane@..."]]
    
    api_headers, api_rows = get_csv_data(api_file)
    # Same thing for API file
    
    # Show what we loaded
    print(f"\nPROD (Master): {len(prod_rows)} rows, {len(prod_headers)} columns")
    print(f"API (Validate): {len(api_rows)} rows, {len(api_headers)} columns")
    # len() = count how many items in a list
    
    # Detect or prompt for key column (prioritize PROD headers)
    key_col = auto_detect_key_column(prod_headers)
    # Call function from Section 3
    # Tries to find "sys_id" or other common ID
    
    if key_col and key_col in api_headers:
        # Check if we found a key AND it exists in API too
        # 'and' = both conditions must be True
        print(f"\n✓ Auto-detected key column: '{key_col}'")
    else:
        # Auto-detect failed, ask the user
        key_col = prompt_for_key_column(prod_headers, api_headers)
        # Call function from Section 4
        print(f"\n✓ Using key column: '{key_col}'")
    
    # Compare columns
    prod_col_set = set(prod_headers)
    api_col_set = set(api_headers)
    # Convert lists to sets for comparison
    # Sets let us do math operations like finding differences
    
    common_columns = prod_col_set & api_col_set
    # & = intersection (columns in BOTH)
    # Example: {"name", "email", "sys_id"}
    
    missing_in_api = prod_col_set - api_col_set
    # - = difference (in PROD but NOT in API)
    # These are columns API should have but doesn't
    # Example: {"phone_number", "address"}
    
    extra_in_api = api_col_set - prod_col_set
    # Columns in API but not in PROD
    # Might be extra fields that shouldn't be there
    # Example: {"debug_field", "temp_value"}
    
    # Create row dictionaries
    prod_dict, prod_empty_keys = create_row_dict(prod_headers, prod_rows, key_col)
    # Call function from Section 5
    # prod_dict = organized by ID for easy lookup
    # prod_empty_keys = problem rows
    
    api_dict, api_empty_keys = create_row_dict(api_headers, api_rows, key_col)
    # Same for API
    
    # Find missing rows (using PROD as baseline)
    prod_keys = set(prod_dict.keys())
    # .keys() = get all IDs from the dictionary
    # prod_keys = {"123", "456", "789"}
    
    api_keys = set(api_dict.keys())
    # api_keys = {"123", "789", "999"}
    
    missing_from_api = prod_keys - api_keys
    # IDs in PROD but not in API
    # missing_from_api = {"456"}
    # This means row with ID "456" is missing from API
    
    extra_in_api_keys = api_keys - prod_keys
    # IDs in API but not in PROD
    # extra_in_api_keys = {"999"}
    # This might be new data or an error
    
    common_keys = prod_keys & api_keys
    # IDs that exist in both files
    # common_keys = {"123", "789"}
    # These are the rows we can compare
    
    # Compare values for common rows (using PROD as source of truth)
    value_differences = []
    # Empty list to store all mismatches we find
    
    matching_rows = []
    # List to store IDs of rows that match perfectly
    
    # Loop through each ID that exists in both files
    for key in common_keys:
        # key = "123" (an ID value)
        
        prod_row = prod_dict[key]['data']
        # Get PROD row with this ID
        # prod_row = ["123", "John", "john@email.com"]
        
        api_row = api_dict[key]['data']
        # Get API row with same ID
        # api_row = ["123", "John", "different@email.com"]
        
        has_difference = False
        # Flag to track if this row has ANY differences
        
        row_differences = []
        # List to store differences found in this row
        
        # Compare each column
        for col in common_columns:
            # Loop through columns that exist in both files
            # col = "email"
            
            if col == key_col:
                continue
                # Skip the key column (we already know it matches)
                # continue = skip to next iteration of loop
            
            prod_idx = prod_headers.index(col)
            # Find position of this column in PROD
            # If col is "email" and headers are ["sys_id", "name", "email"]
            # Then prod_idx = 2
            
            api_idx = api_headers.index(col)
            # Same for API (might be different position)
            
            prod_val = prod_row[prod_idx].strip() if prod_idx < len(prod_row) else ""
            # Get the value from PROD for this column
            # prod_row[2] = "john@email.com"
            # .strip() = remove spaces
            # 'if prod_idx < len(prod_row) else ""' = safety check
            # If row is too short, use empty string instead of crashing
            
            api_val = api_row[api_idx].strip() if api_idx < len(api_row) else ""
            # Same for API
            # api_val = "different@email.com"
            
            if api_val != prod_val:
                # != means "not equal"
                # Check if values are different
                
                has_difference = True
                # Mark that this row has issues
                
                row_differences.append({
                    'key': key,
                    'column': col,
                    'prod_value': prod_val,
                    'api_value': api_val,
                    'issue': 'Missing' if not api_val and prod_val else 'Different'
                })
                # .append() = add to list
                # Store details about this difference
                # {} = dictionary with multiple pieces of info
                # 'issue' = if API value is empty, it's "Missing"
                #           if API has wrong value, it's "Different"
                # 'if not api_val and prod_val' = API is empty AND PROD has value
        
        if has_difference:
            # This row had at least one mismatch
            value_differences.extend(row_differences)
            # .extend() = add all items from row_differences to value_differences
            # Like combining two lists
        else:
            # This row matched perfectly
            matching_rows.append(key)
            # Add this ID to our list of perfect matches
    
    # Calculate API data quality metrics
    total_prod_rows = len(prod_dict)
    # How many rows in PROD (our baseline)
    
    api_completeness = (len(common_keys) / total_prod_rows * 100) if total_prod_rows > 0 else 0
    # Completeness = what % of PROD rows exist in API
    # len(common_keys) = rows in both files
    # / total_prod_rows = divide to get fraction
    # * 100 = convert to percentage
    # 'if total_prod_rows > 0 else 0' = prevent division by zero error
    # Example: 85 / 100 * 100 = 85%
    
    api_accuracy = (len(matching_rows) / len(common_keys) * 100) if common_keys else 0
    # Accuracy = of the rows that exist, what % are 100% correct
    # len(matching_rows) = perfectly matching rows
    # len(common_keys) = all rows in both files
    # Example: 70 / 85 * 100 = 82.4%
    
    # Print summary to console
    print("\n" + "="*70)
    print("VALIDATION SUMMARY")
    print("="*70)
    
    print(f"\n📊 COLUMN COMPARISON:")
    print(f"   ✓ Common columns: {len(common_columns)}")
    print(f"   ⚠ Missing in API (present in PROD): {len(missing_in_api)}")
    print(f"   ⚡ Extra in API (not in PROD): {len(extra_in_api)}")
    
    print(f"\n📊 ROW COMPARISON (PROD as baseline):")
    print(f"   Total PROD rows: {total_prod_rows}")
    print(f"   ✓ Found in API: {len(common_keys)}")
    print(f"   ⚠ Missing from API: {len(missing_from_api)}")
    print(f"   ⚡ Extra in API: {len(extra_in_api_keys)}")
    
    print(f"\n📊 DATA QUALITY:")
    print(f"   ✓ Rows matching PROD exactly: {len(matching_rows)}")
    print(f"   ⚠ Rows with differences: {len(common_keys) - len(matching_rows)}")
    print(f"   ⚠ Total value mismatches: {len(value_differences)}")
    print(f"   ⚠ API rows with empty keys: {len(api_empty_keys)}")
    if prod_empty_keys:
        print(f"   ⚠ PROD rows with empty keys: {len(prod_empty_keys)}")
    
    print(f"\n📊 API COMPLETENESS: {api_completeness:.1f}% of PROD data")
    # .1f = format number to 1 decimal place
    # 85.345 becomes "85.3%"
    
    print(f"📊 API ACCURACY: {api_accuracy:.1f}% of matched rows are correct")
    
    # Create Excel report
    create_excel_report(
        output_excel, key_col, prod_headers, api_headers,
        common_columns, missing_in_api, extra_in_api,
        missing_from_api, extra_in_api_keys, prod_dict, api_dict,
        value_differences, matching_rows, api_empty_keys, prod_empty_keys,
        api_completeness, api_accuracy
    )
    # Call the function to create Excel (defined in next section)
    # Pass all the data we collected
    
    print(f"\n✓ Excel validation report created: {output_excel}")
    print("="*70)


# ============================================================================
# SECTION 7: CREATE EXCEL REPORT
# ============================================================================
# This section is very long - it creates all the Excel sheets
# I'll explain the key concepts

def create_excel_report(filename, key_col, prod_headers, api_headers,
                       common_columns, missing_in_api, extra_in_api,
                       missing_from_api, extra_in_api_keys, prod_dict, api_dict,
                       value_differences, matching_rows, api_empty_keys, prod_empty_keys,
                       api_completeness, api_accuracy):
    """Create Excel workbook with validation results."""
    
    wb = openpyxl.Workbook()
    # Create new Excel file
    # wb = workbook (the whole Excel file)
    
    # Define color styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    # Blue background for headers
    # "4472C4" = hex color code (blue)
    
    header_font = Font(bold=True, color="FFFFFF")
    # White, bold text for headers
    
    error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    # Red background for errors
    
    warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    # Yellow background for warnings
    
    success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    # Green background for success
    
    # ========================================================================
    # Sheet 1: Summary
    # ========================================================================
    
    ws_summary = wb.active
    # wb.active = the first/default sheet in the workbook
    
    ws_summary.title = "Summary"
    # Rename the sheet tab to "Summary"
    
    # Create summary data as list of lists
    # Each inner list becomes one row in Excel
    summary_data = [
        ["API VALIDATION REPORT - PROD as Master Source", ""],
        ["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        # strftime = format date/time as string
        # %Y = 4-digit year, %m = month, %d = day, %H = hour, %M = minute, %S = second
        ["Key Column Used", key_col],
        ["", ""],  # Blank row
        ["COLUMN COMPARISON", ""],
        ["Total PROD Columns", len(prod_headers)],
        ["Total API Columns", len(api_headers)],
        # ... more rows ...
    ]
    
    # Add each row to the worksheet
    for row_data in summary_data:
        ws_summary.append(row_data)
        # .append() adds a new row to the sheet
    
    # Format the sheet
    ws_summary.column_dimensions['A'].width = 35
    # Set column A width to 35 characters
    
    ws_summary.column_dimensions['B'].width = 25
    # Set column B width
    
    # Make certain rows bold (headers)
    for row in [1, 5, 12, 19, 26]:
        ws_summary[f'A{row}'].font = Font(bold=True, size=12)
        # ws_summary['A1'] = cell A1
        # .font = change font style
    
    # Color code the quality scores
    completeness_row = 26
    if api_completeness >= 95:
        ws_summary[f'B{completeness_row}'].fill = success_fill  # Green
    elif api_completeness >= 80:
        ws_summary[f'B{completeness_row}'].fill = warning_fill  # Yellow
    else:
        ws_summary[f'B{completeness_row}'].fill = error_fill    # Red
    
    # Same for accuracy
    accuracy_row = 27
    if api_accuracy >= 95:
        ws_summary[f'B{accuracy_row}'].fill = success_fill
    elif api_accuracy >= 80:
        ws_summary[f'B{accuracy_row}'].fill = warning_fill
    else:
        ws_summary[f'B{accuracy_row}'].fill = error_fill
    
    # ========================================================================
    # Sheet 2: Column Comparison
    # ========================================================================
    
    ws_columns = wb.create_sheet("Column Comparison")
    # Create new sheet named "Column Comparison"
    
    ws_columns.append(["Status", "Column Name", "Notes"])
    # Add header row
    
    # Format header row
    for col in ws_columns[1]:
        # ws_columns[1] = first row (headers)
        col.fill = header_fill    # Blue background
        col.font = header_font    # White, bold text
    
    # Add common columns
    for col in sorted(common_columns):
        ws_columns.append(["✓ Common", col, "Present in both"])
    
    # Add missing columns (with red highlighting)
    for col in sorted(missing_in_api):
        row = ws_columns.max_row + 1
        # max_row = current last row number
        # + 1 = next row
        ws_columns.append(["⚠ MISSING IN API", col, "Required from PROD"])
        ws_columns[f'A{row}'].fill = error_fill  # Red
        ws_columns[f'B{row}'].fill = error_fill  # Red
    
    # Add extra columns
    for col in sorted(extra_in_api):
        ws_columns.append(["⚡ Extra in API", col, "Not in PROD"])
    
    # Set column widths
    ws_columns.column_dimensions['A'].width = 20
    ws_columns.column_dimensions['B'].width = 40
    ws_columns.column_dimensions['C'].width = 30
    
    # ========================================================================
    # Sheet 3: Missing Rows from API
    # ========================================================================
    
    ws_missing = wb.create_sheet("Missing from API")
    # Create new sheet for rows that should be in API but aren't
    
    ws_missing.append([key_col, "Status", "PROD Data Preview"])
    # Add header row
    
    # Format headers
    for col in ws_missing[1]:
        col.fill = header_fill
        col.font = header_font
    
    # Add each missing row
    for key in sorted(missing_from_api):
        # key = the ID that's missing (like "456")
        
        row_data = prod_dict[key]['data']
        # Get the full row from PROD
        # row_data = ["456", "Jane", "jane@email.com", "555-1234"]
        
        # Create preview of first 4 columns
        info = " | ".join([f"{prod_headers[i]}: {row_data[i]}" for i in range(min(4, len(row_data))) if i < len(prod_headers)])
        # This complex line creates a preview string
        # range(min(4, len(row_data))) = loop through first 4 columns
        # f"{prod_headers[i]}: {row_data[i]}" = "name: Jane"
        # " | ".join() = combine with " | " between each
        # Result: "sys_id: 456 | name: Jane | email: jane@email.com | phone: 555-1234"
        
        row_num = ws_missing.max_row + 1
        # Get next row number
        
        ws_missing.append([key, "⚠ MISSING", info])
        # Add row to Excel
        
        # Highlight in red (this is critical - row is missing!)
        ws_missing[f'A{row_num}'].fill = error_fill
        ws_missing[f'B{row_num}'].fill = error_fill
    
    # Set column widths
    ws_missing.column_dimensions['A'].width = 30
    ws_missing.column_dimensions['B'].width = 15
    ws_missing.column_dimensions['C'].width = 80
    
    # ========================================================================
    # Sheet 4: Value Differences
    # ========================================================================
    
    ws_diff = wb.create_sheet("Value Differences")
    # Sheet for rows that exist but have wrong values
    
    ws_diff.append([key_col, "Column", "Issue Type", "PROD Value (Correct)", "API Value (Incorrect)"])
    # Headers - note we label PROD as "Correct"
    
    # Format headers
    for col in ws_diff[1]:
        col.fill = header_fill
        col.font = header_font
    
    # Add each value difference
    for diff in value_differences:
        # diff = dictionary with info about one mismatch
        # diff = {'key': '123', 'column': 'email', 'issue': 'Different',
        #         'prod_value': 'john@email.com', 'api_value': 'wrong@email.com'}
        
        row_num = ws_diff.max_row + 1
        
        ws_diff.append([
            diff['key'],           # The row ID
            diff['column'],        # Which column has the issue
            diff['issue'],         # "Missing" or "Different"
            diff['prod_value'],    # What it should be (from PROD)
            diff['api_value']      # What API has (wrong)
        ])
        
        # Color code based on issue type
        if diff['issue'] == 'Missing':
            # API has empty value where PROD has data - critical error
            ws_diff[f'C{row_num}'].fill = error_fill  # Red
            ws_diff[f'E{row_num}'].fill = error_fill  # Red
        else:
            # API has different (but not empty) value - warning
            ws_diff[f'C{row_num}'].fill = warning_fill  # Yellow
    
    # Set column widths
    ws_diff.column_dimensions['A'].width = 25
    ws_diff.column_dimensions['B'].width = 25
    ws_diff.column_dimensions['C'].width = 15
    ws_diff.column_dimensions['D'].width = 35
    ws_diff.column_dimensions['E'].width = 35
    
    # ========================================================================
    # Sheet 5: Extra in API
    # ========================================================================
    
    ws_extra = wb.create_sheet("Extra in API")
    # Sheet for rows in API that don't exist in PROD
    # These might be new data or errors
    
    ws_extra.append([key_col, "Status", "API Data Preview", "Notes"])
    
    # Format headers
    for col in ws_extra[1]:
        col.fill = header_fill
        col.font = header_font
    
    # Add each extra row
    for key in sorted(extra_in_api_keys):
        row_data = api_dict[key]['data']
        # Get row from API
        
        # Create preview
        info = " | ".join([f"{api_headers[i]}: {row_data[i]}" for i in range(min(4, len(row_data))) if i < len(api_headers)])
        
        ws_extra.append([key, "Extra", info, "Not in PROD - may be new or error"])
        # We don't color these red because they might be legitimate new data
    
    # Set column widths
    ws_extra.column_dimensions['A'].width = 30
    ws_extra.column_dimensions['B'].width = 15
    ws_extra.column_dimensions['C'].width = 80
    ws_extra.column_dimensions['D'].width = 30
    
    # ========================================================================
    # Sheet 6: Data Issues (only if there are empty key problems)
    # ========================================================================
    
    if api_empty_keys or prod_empty_keys:
        # Only create this sheet if there are rows with empty keys
        
        ws_issues = wb.create_sheet("Data Issues")
        ws_issues.append(["Issue Type", "Source", "Row Number", "Data Preview"])
        
        # Format headers
        for col in ws_issues[1]:
            col.fill = header_fill
            col.font = header_font
        
        # Add API rows with empty keys
        for item in api_empty_keys:
            # item = {'row_num': 5, 'data': ['', 'John', 'john@email.com']}
            
            # Create preview
            info = " | ".join([f"{api_headers[i]}: {item['data'][i]}" for i in range(min(4, len(item['data']))) if i < len(api_headers)])
            
            row_num = ws_issues.max_row + 1
            ws_issues.append([f"Empty {key_col}", "API", item['row_num'], info])
            # Show which row in the CSV file has the problem
            
            ws_issues[f'A{row_num}'].fill = error_fill  # Red - this is bad
        
        # Add PROD rows with empty keys
        for item in prod_empty_keys:
            info = " | ".join([f"{prod_headers[i]}: {item['data'][i]}" for i in range(min(4, len(item['data']))) if i < len(prod_headers)])
            
            row_num = ws_issues.max_row + 1
            ws_issues.append([f"Empty {key_col}", "PROD", item['row_num'], info])
            
            ws_issues[f'A{row_num}'].fill = warning_fill  # Yellow - concerning but PROD might have reason
        
        # Set column widths
        ws_issues.column_dimensions['A'].width = 20
        ws_issues.column_dimensions['B'].width = 15
        ws_issues.column_dimensions['C'].width = 15
        ws_issues.column_dimensions['D'].width = 80
    
    # ========================================================================
    # Save the Excel file
    # ========================================================================
    
    wb.save(filename)
    # Save all sheets to the Excel file
    # filename = "api_validation_report_20250102_143022.xlsx"


# ============================================================================
# SECTION 8: MAIN PROGRAM EXECUTION
# ============================================================================
# This is where the program actually starts running

if __name__ == "__main__":
    # This special check means "only run this if we're running the script directly"
    # If someone imports this script into another program, this part won't run
    # It's a Python best practice
    
    # ========================================================================
    # CONFIGURE YOUR FILE PATHS HERE
    # ========================================================================
    
    PROD_CSV = "prod_data.csv"  # CHANGE THIS to your PROD file path
    # Master source of truth
    # Example: "C:/data/servicenow_prod.csv"
    # or: r"Y:\My Documents\Data\prod.csv"
    
    API_CSV = "api_data.csv"    # CHANGE THIS to your API file path
    # Data to validate against PROD
    # Example: "C:/data/api_export.csv"
    
    OUTPUT_EXCEL = f"api_validation_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    # Output filename with timestamp
    # f"..." = f-string lets us insert variables
    # datetime.now().strftime('%Y%m%d_%H%M%S') creates timestamp like "20250102_143022"
    # Final name: "api_validation_report_20250102_143022.xlsx"
    # This prevents overwriting old reports
    
    # ========================================================================
    # RUN THE COMPARISON
    # ========================================================================
    
    print("\nStarting API validation against PROD (Master Source)...")
    # \n = start with blank line for readability
    
    print(f"PROD CSV (Master): {PROD_CSV}")
    print(f"API CSV (Validate): {API_CSV}")
    print(f"Output: {OUTPUT_EXCEL}")
    # Show user what files we're processing
    
    compare_csv_files(PROD_CSV, API_CSV, OUTPUT_EXCEL)
    # Call the main function - this does everything!
    # This function will:
    # 1. Read both CSV files
    # 2. Find key column
    # 3. Compare everything
    # 4. Create Excel report
    # 5. Print summary to console


# ============================================================================
# HOW TO USE THIS SCRIPT
# ============================================================================
"""
STEP 1: Install required library
   Open terminal/command prompt and run:
   pip install openpyxl

STEP 2: Update file paths
   Edit lines with PROD_CSV and API_CSV to point to your files
   
STEP 3: Run the script
   In VS Code: Press F5 or click Run button
   In terminal: python script_name.py

STEP 4: Check the output
   - Console will show summary
   - Excel file will be created in same folder as script
   - Excel has 6 sheets with detailed findings

STEP 5: Review Excel report
   - Summary tab: Overall quality scores
   - Missing from API: Critical - rows API doesn't have
   - Value Differences: Where API data is wrong
   - Other tabs: Additional details

KEY CONCEPTS EXPLAINED:
=======================

1. LISTS [ ]
   - Ordered collection of items
   - Example: ["apple", "banana", "cherry"]
   - Access by position: fruits[0] = "apple"
   - Can have duplicates

2. DICTIONARIES { }
   - Key-value pairs (like a phonebook)
   - Example: {"name": "John", "age": 30}
   - Access by key: person["name"] = "John"
   - Keys must be unique

3. SETS { }
   - Unordered collection of unique items
   - Example: {"apple", "banana", "cherry"}
   - Good for comparisons (intersection, difference)
   - No duplicates allowed

4. FOR LOOPS
   for item in list:
       # Do something with item
   - Repeats code for each item in a collection

5. IF STATEMENTS
   if condition:
       # Do this if True
   else:
       # Do this if False
   - Makes decisions in code

6. FUNCTIONS
   def function_name(parameters):
       # Code here
       return result
   - Reusable blocks of code
   - Take inputs (parameters)
   - Give outputs (return value)

7. F-STRINGS
   f"Hello {name}"
   - Put variables inside strings
   - Example: f"Score: {score}%" → "Score: 85%"

8. METHODS
   string.strip()  # Remove spaces
   list.append()   # Add to list
   dict.keys()     # Get all keys
   - Functions that belong to objects
   - Called with dot notation

9. COMPARISON OPERATORS
   == equal to
   != not equal to
   <  less than
   >  greater than
   <= less than or equal
   >= greater than or equal

10. LOGICAL OPERATORS
    and = both must be True
    or  = at least one must be True
    not = opposite of True/False

COMMON PYTHON PATTERNS IN THIS SCRIPT:
======================================

Pattern: Loop with enumerate
   for i, item in enumerate(list, start=1):
   # i = counter (1, 2, 3...)
   # item = actual item from list

Pattern: List comprehension
   [x.strip() for x in list]
   # Apply strip() to each item, create new list

Pattern: Dictionary comprehension
   {key: value for key, value in items}
   # Create dictionary from items

Pattern: Conditional expression (ternary operator)
   value = a if condition else b
   # If condition is True, use a, otherwise use b

Pattern: Try-except
   try:
       # Try this code
   except ErrorType:
       # If error happens, do this

Pattern: File handling with 'with'
   with open(file) as f:
       # Work with file
   # File automatically closes

Pattern: String formatting
   f"{variable:.1f}%"
   # .1f = one decimal place
   # 85.345 becomes "85.3%"

TROUBLESHOOTING:
===============

Error: "No module named 'openpyxl'"
Solution: pip install openpyxl

Error: "File not found"
Solution: Check file paths are correct

Error: "Permission denied"
Solution: Close Excel file if it's open

Error: "Encoding error"
Solution: Script tries multiple encodings automatically

TERMINOLOGY
==========================

Data Validation: Checking if API data matches expected standards (PROD)

Master Source: PROD is the "source of truth" - the correct data

Completeness: What % of expected data exists
   Example: "API is 85% complete"

Accuracy: What % of existing data is correct
   Example: "API is 90% accurate"

Key Column: Unique identifier used to match rows between files
   Example: sys_id, employee_id, ticket_number

Missing Rows: Records that should exist but don't

Value Mismatches: Records exist but have incorrect data

Data Quality Metrics: Measurements of how good the data is
"""
